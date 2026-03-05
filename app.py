import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
import math
import zipfile
import os
import re

# ---------------------------
# Streamlit UI
# ---------------------------
st.set_page_config(page_title="QA Assignment Tool", layout="wide")
st.title("QA Assignment Tool 📊")
st.write("Even split with brand integrity. Priorities: 1) Perfect balance, 2) Pre-assignments, 3) Keep brands together")

# ---------------------------
# Helpers
# ---------------------------

def get_header_map(worksheet):
    """Build a dictionary mapping header names to column indices (1-based)."""
    header_map = {}
    for col_idx, cell in enumerate(worksheet[1], start=1):
        if cell.value:
            header_map[cell.value.strip()] = col_idx
            header_map[cell.value.strip().lower()] = col_idx
    return header_map


def get_col_index(header_map, *possible_names):
    """Get column index from header map, trying multiple possible header names."""
    for name in possible_names:
        if name in header_map:
            return header_map[name]
        if name.lower() in header_map:
            return header_map[name.lower()]
    raise KeyError(f"Could not find column with any of these headers: {possible_names}")


def title_or_none(val):
    return val.strip().title() if isinstance(val, str) and val.strip() else None


def calculate_exact_targets(active_members, total_products, member_limits):
    """
    Calculate EXACT targets for perfect distribution, respecting member limits.

    When a member's limit is below their fair share, that capacity is redistributed
    to other members who can take more. This iterates until stable.
    """
    remaining = total_products
    targets = {m: 0 for m in active_members}
    locked = set()

    max_iterations = 100
    iteration = 0

    while iteration < max_iterations:
        iteration += 1
        unlocked = [m for m in active_members if m not in locked]

        if not unlocked:
            break

        per_person = remaining // len(unlocked)
        remainder = remaining % len(unlocked)

        changed = False
        temp_assignments = {}

        for i, m in enumerate(unlocked):
            fair_share = per_person + (1 if i < remainder else 0)
            limit = member_limits.get(m, 999)

            if limit < fair_share:
                temp_assignments[m] = limit
                locked.add(m)
                changed = True
            else:
                temp_assignments[m] = fair_share

        for m, count in temp_assignments.items():
            targets[m] = count

        remaining = total_products - sum(targets[m] for m in locked)

        if not changed:
            break

    return targets


def get_member_furthest_from_target(active_members, counts, targets, required_space=1):
    """Find member with most room to their target who has required_space available."""
    candidates = []
    for m in active_members:
        room = targets[m] - counts[m]
        if room >= required_space:
            candidates.append((m, room))

    if not candidates:
        return None

    candidates.sort(key=lambda x: -x[1])
    return candidates[0][0]


def get_members_with_room(active_members, counts, targets):
    """Get list of members who are still below their target, sorted by most room."""
    members = [(m, targets[m] - counts[m]) for m in active_members if counts[m] < targets[m]]
    members.sort(key=lambda x: -x[1])
    return [m for m, _ in members]


def clean_output_xlsx(output_path):
    """
    Post-process the saved xlsx to remove external links and stale calcChain,
    which cause Excel's repair/recovery prompt on open.
    """
    clean_path = output_path.replace(".xlsx", "_clean.xlsx")

    skip_files = {
        'xl/calcChain.xml',
        'xl/externalLinks/externalLink1.xml',
        'xl/externalLinks/externalLink2.xml',
        'xl/externalLinks/_rels/externalLink1.xml.rels',
        'xl/externalLinks/_rels/externalLink2.xml.rels',
    }

    with zipfile.ZipFile(output_path, 'r') as zin:
        with zipfile.ZipFile(clean_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename in skip_files:
                    continue
                data = zin.read(item.filename)
                # Scrub externalReferences block from workbook.xml
                if item.filename == 'xl/workbook.xml':
                    data = re.sub(rb'<externalReference[^/]*/>', b'', data)
                    data = re.sub(
                        rb'<externalReferences>.*?</externalReferences>',
                        b'',
                        data,
                        flags=re.DOTALL
                    )
                # Scrub externalLink relationships from workbook.xml.rels
                if item.filename == 'xl/_rels/workbook.xml.rels':
                    data = re.sub(rb'<Relationship[^>]*externalLink[^>]*/>', b'', data)
                zout.writestr(item, data)

    os.replace(clean_path, output_path)


# ---------------------------
# File upload
# ---------------------------
uploaded_file = st.file_uploader("Upload QA Template", type=["xlsx"])
if not uploaded_file:
    st.info("Please upload an Excel (.xlsx) file containing 'QA' and 'Assignments' sheets.")
    st.stop()

temp_file_path = f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
with open(temp_file_path, "wb") as f:
    f.write(uploaded_file.getbuffer())

wb = load_workbook(temp_file_path, keep_links=False)
if "QA" not in wb.sheetnames or "Assignments" not in wb.sheetnames:
    st.error("Excel file must contain 'QA' and 'Assignments' sheets.")
    st.stop()

qa_ws = wb["QA"]
assignments_ws = wb["Assignments"]

# ---------------------------
# Build header maps
# ---------------------------
qa_headers = get_header_map(qa_ws)
assignments_headers = get_header_map(assignments_ws)

try:
    COL_ASSIGNED = get_col_index(qa_headers, "Assigned", "assigned", "ASSIGNED")
    COL_PIM_PARENT_ID = get_col_index(qa_headers, "Pim Parent ID", "pim parent id", "PIM Parent ID")
    COL_BRAND = get_col_index(qa_headers, "Brand", "brand", "BRAND")
    COL_BT_IMAGE_DATE = get_col_index(qa_headers, "Bt Image Date", "bt image date", "BT Image Date",
                                       "Enrichment QA Date", "enrichment qa date")
except KeyError as e:
    st.error(f"Missing required column in QA sheet: {e}")
    st.stop()

try:
    COL_ASSIGN_BRAND = get_col_index(assignments_headers, "BRAND", "Brand", "brand")
    COL_ASSIGN_QAER = get_col_index(assignments_headers, "Qaer", "qaer", "QAER", "QA", "Member", "member")
except KeyError as e:
    st.error(f"Missing required column in Assignments sheet: {e}")
    st.stop()

with st.expander("📋 Detected Column Mappings"):
    st.write("**QA Sheet:**")
    st.write(f"- Assigned: Column {get_column_letter(COL_ASSIGNED)}")
    st.write(f"- Pim Parent ID: Column {get_column_letter(COL_PIM_PARENT_ID)}")
    st.write(f"- Brand: Column {get_column_letter(COL_BRAND)}")
    st.write(f"- BT Image Date: Column {get_column_letter(COL_BT_IMAGE_DATE)}")
    st.write("**Assignments Sheet:**")
    st.write(f"- Brand: Column {get_column_letter(COL_ASSIGN_BRAND)}")
    st.write(f"- Qaer: Column {get_column_letter(COL_ASSIGN_QAER)}")

# ---------------------------
# Options
# ---------------------------
st.subheader("⚙️ Configuration")

backlog_mode = st.checkbox("Backlog mode (sort by earliest BT Image date)", value=False)

st.write("Enter active members today (e.g: Ross:100, Phoebe:80, Monica)")
working_input = st.text_input("Active members")

if not working_input:
    st.error("Please enter at least one active member.")
    st.stop()

# Parse active members
active_members = []
member_limits = {}
for part in working_input.split(","):
    part = part.strip()
    if not part:
        continue
    if ":" in part:
        name, limit = part.split(":", 1)
        name = name.strip().title()
        try:
            lim = int(limit.strip())
        except:
            lim = 999
        active_members.append(name)
        member_limits[name] = lim
    else:
        active_members.append(part.strip().title())

for m in active_members:
    if m not in member_limits:
        member_limits[m] = 999

# Preassignments from Assignments sheet
brand_to_member = {}
for row in assignments_ws.iter_rows(min_row=2, values_only=True):
    brand = row[COL_ASSIGN_BRAND - 1] if len(row) >= COL_ASSIGN_BRAND else None
    member = row[COL_ASSIGN_QAER - 1] if len(row) >= COL_ASSIGN_QAER else None
    if brand and member:
        brand_to_member[title_or_none(brand)] = title_or_none(member)

# Show pre-assignments
if brand_to_member:
    with st.expander(f"📌 Pre-assigned Brands ({len(brand_to_member)})"):
        for brand, member in sorted(brand_to_member.items()):
            status = "✅" if member in active_members else "⚠️ (not active today)"
            st.write(f"- {brand} → {member} {status}")

# Build brand blocks
brand_blocks = defaultdict(list)
row_brand_order = []
row_to_brand = {}

for i, row in enumerate(qa_ws.iter_rows(min_row=2, values_only=True), start=2):
    pim_parent_id = row[COL_PIM_PARENT_ID - 1] if len(row) >= COL_PIM_PARENT_ID else None
    brand = row[COL_BRAND - 1] if len(row) >= COL_BRAND else None

    if pim_parent_id is not None and str(pim_parent_id).strip():
        btitle = title_or_none(brand) if brand else "No Brand"
        row_to_brand[i] = btitle
        if btitle not in brand_blocks:
            row_brand_order.append(btitle)
        brand_blocks[btitle].append(i)

if backlog_mode:
    date_col_letter = get_column_letter(COL_BT_IMAGE_DATE)

    def row_date(row_idx):
        try:
            val = qa_ws[f"{date_col_letter}{row_idx}"].value
        except:
            val = None
        return val if isinstance(val, datetime) else datetime.max

    for b in brand_blocks:
        brand_blocks[b].sort(key=row_date)

# Build blocks list with pre-assignment info
blocks = []
for b in row_brand_order:
    pre_member = brand_to_member.get(b)
    is_preassigned = pre_member is not None and pre_member in active_members
    blocks.append({
        'brand': b,
        'rows': brand_blocks[b].copy(),
        'size': len(brand_blocks[b]),
        'preassigned_to': pre_member if is_preassigned else None
    })

# Sort: pre-assigned first, then by size (smallest first)
blocks.sort(key=lambda x: (0 if x['preassigned_to'] else 1, x['size']))

# Calculate totals and EXACT targets
total_products = sum(b['size'] for b in blocks)
targets = calculate_exact_targets(active_members, total_products, member_limits)

# Check total capacity
total_capacity = sum(targets.values())
if total_capacity < total_products:
    shortfall = total_products - total_capacity
    st.warning(f"⚠️ Total capacity ({total_capacity}) is less than products ({total_products}). {shortfall} will go to backlog.")

# Display targets
st.write("📊 **Exact Targets for Perfect Split:**")
target_cols = st.columns(len(active_members))
for i, m in enumerate(active_members):
    with target_cols[i]:
        limit_info = f" (limit: {member_limits[m]})" if member_limits[m] < 999 else ""
        st.metric(m, f"{targets[m]} products", delta=limit_info if limit_info else None)

st.divider()

# ---------------------------
# PERFECT EVEN SPLIT ALGORITHM
# ---------------------------
counts = {m: 0 for m in active_members}
assignments = {m: [] for m in active_members}
brand_assignments_log = []
backlog_rows = []
assigned_col_letter = get_column_letter(COL_ASSIGNED)

for block in blocks:
    brand = block['brand']
    rows = block['rows'].copy()
    preassigned_to = block['preassigned_to']

    if not rows:
        continue

    # If pre-assigned, try to give to that member first
    if preassigned_to:
        room = targets[preassigned_to] - counts[preassigned_to]
        if room > 0:
            take = min(room, len(rows))
            taken_rows = rows[:take]
            rows = rows[take:]

            for r in taken_rows:
                qa_ws[f"{assigned_col_letter}{r}"].value = preassigned_to
            assignments[preassigned_to].extend(taken_rows)
            counts[preassigned_to] += len(taken_rows)

            brand_assignments_log.append({
                'brand': brand,
                'size': len(taken_rows),
                'member': preassigned_to,
                'preassigned': True,
                'split': len(rows) > 0
            })

    # Distribute remaining rows
    while rows:
        members_with_room = get_members_with_room(active_members, counts, targets)

        if not members_with_room:
            for r in rows:
                qa_ws[f"{assigned_col_letter}{r}"].value = "Backlog"
                backlog_rows.append(r)
            break

        brand_size = len(rows)
        best_single_member = None

        for m in members_with_room:
            room = targets[m] - counts[m]
            if room >= brand_size:
                best_single_member = m
                break

        if best_single_member:
            for r in rows:
                qa_ws[f"{assigned_col_letter}{r}"].value = best_single_member
            assignments[best_single_member].extend(rows)
            counts[best_single_member] += len(rows)

            brand_assignments_log.append({
                'brand': brand,
                'size': len(rows),
                'member': best_single_member,
                'preassigned': False,
                'split': preassigned_to is not None
            })
            rows = []
        else:
            for m in members_with_room:
                if not rows:
                    break

                room = targets[m] - counts[m]
                if room <= 0:
                    continue

                take = min(room, len(rows))
                taken_rows = rows[:take]
                rows = rows[take:]

                for r in taken_rows:
                    qa_ws[f"{assigned_col_letter}{r}"].value = m
                assignments[m].extend(taken_rows)
                counts[m] += len(taken_rows)

                brand_assignments_log.append({
                    'brand': brand,
                    'size': len(taken_rows),
                    'member': m,
                    'preassigned': False,
                    'split': True
                })

# ---------------------------
# FINAL BALANCE CHECK
# ---------------------------
final_adjustments = 0
max_iterations = 1000
iteration = 0

while iteration < max_iterations:
    iteration += 1

    over = [(m, counts[m] - targets[m]) for m in active_members if counts[m] > targets[m]]
    under = [(m, targets[m] - counts[m]) for m in active_members if counts[m] < targets[m]]

    if not over or not under:
        break

    over.sort(key=lambda x: -x[1])
    under.sort(key=lambda x: -x[1])

    from_member = over[0][0]
    to_member = under[0][0]

    if not assignments[from_member]:
        break

    row_to_move = assignments[from_member].pop()
    assignments[to_member].append(row_to_move)
    counts[from_member] -= 1
    counts[to_member] += 1
    qa_ws[f"{assigned_col_letter}{row_to_move}"].value = to_member
    final_adjustments += 1

# ---------------------------
# Results Summary
# ---------------------------
st.subheader("📈 Final Distribution")

result_cols = st.columns(len(active_members))
for i, m in enumerate(active_members):
    with result_cols[i]:
        diff = counts[m] - targets[m]
        if diff == 0:
            st.success(f"**{m}**: {counts[m]} ✓")
        elif diff > 0:
            st.warning(f"**{m}**: {counts[m]} (+{diff})")
        else:
            st.error(f"**{m}**: {counts[m]} ({diff})")

if backlog_rows:
    st.warning(f"📦 **Backlog**: {len(backlog_rows)} products")

if final_adjustments > 0:
    st.info(f"🔄 Made {final_adjustments} final adjustments for perfect balance")

# ---------------------------
# Save, clean, and offer download
# ---------------------------
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
output_path = f"QA_Assignment_{timestamp}.xlsx"
wb.save(output_path)

# Remove external links and stale calcChain to prevent Excel repair prompt
clean_output_xlsx(output_path)

st.success("✅ Assignment complete!")

with open(output_path, "rb") as f:
    st.download_button(
        label="📥 Download Assigned Excel",
        data=f,
        file_name=output_path,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
